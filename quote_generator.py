from openpyxl import load_workbook
from random import randint
from time import sleep
import warnings

# Helps out openpyxl.
warnings.simplefilter("ignore")

# Function for removing newline characters.
def strip_string(string):
	string = string.replace('\n','')
	string = string.rstrip()
	return string
	
	
lastquote_filepath = 'lastquote.txt'
padding = 20
sleep_sec = 0


# Load in workbook.
workbook_name = 'data.xlsx'
wb = load_workbook(workbook_name)
quotes_sheet = wb['quotes']
used_numbers_sheet = wb['used_numbers']

# Calculate number of quotes (rows in quotes sheet).
num_quotes = quotes_sheet.max_row

# Generate used_numbers list.
used_numbers = []
for number in used_numbers_sheet['A']:
	if number.value is not None:
		used_numbers.append(number.value)

# Check if all quotes have been used.
if num_quotes == len(used_numbers):
	used_numbers = []
	used_numbers_sheet.delete_cols(1)
	wb.save(filename = workbook_name)

while True:
	# Pick a random number between 1 and number of quotes.
	random_number = randint(1, num_quotes)
	
	# Check to see if this quote has been picked.	
	if random_number in used_numbers:
		pass
	else:
		# Write number to list.
		used_numbers_sheet[f'A{len(used_numbers) + 1}'] = random_number
		used_numbers.append(random_number)
		
		# Retrieve quote.
		quote = quotes_sheet[f'A{random_number}'].value
		
		wb.save(filename = workbook_name)

		# Write new quote to .txt with space padding for scrolling.
		with open(lastquote_filepath, "w", encoding='utf-8') as f:
			f.write(quote + (' ' * padding))
			f.close()
		break

# Feedback:

# Print quote.
print()
print(quote)
print()
print(f'{len(used_numbers)} / {num_quotes} used.')
print()


# Sleep.
sleep(sleep_sec)

input("Press any key to exit...")
